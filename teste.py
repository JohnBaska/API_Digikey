import requests
import json
import time
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import xlsxwriter
from urllib.parse import quote, urlparse, parse_qs

class API():

    def __init__ (self):
        self.partnumbers = []

        # código de acesso da pagina localhost
        self.code = 'OPA43qkZ'

        # arquivo json que armazena as informações de acesso da digikey
        self.filename = 'digikey_token.json'

        # alimento o token
        self.load_token_from_file()

        op = 1

        # Menu Principal do programa
        while (op != 0):
            print ("(0) Sair")
            print ("(1) Gerar token")
            print ("(2) Atualizar token")
            print ("(3) Alimentar o dados.json") 
            print ("(4) Alimentar a planilha de saída")

            op = int(input("Escolha: "))

            if op == 1:
                self.get_access_token()
            elif op == 2:
                self.get_refresh_token()
            elif op == 3:
                self.sheet = input("Caminho da panilha")
                self.sheet = "entradas.xlsx"

                self.get_dates_sheet()

                self.data = self.get_product_details()
            elif op == 4:
                self.filling_out_spreadsheet()
                self.style_sheet()
            elif op == 0:
                print ("Você saiu do código")
            else:
                print('opcao invalida')

    ################ Alimenta o token ################

    def load_token_from_file(self):
        with open(self.filename, 'r') as arquivo:
            self.token = json.load(arquivo)

        if self.token != False:
            print('\033[32mToken load SUCCESS.\033[0m')
        else:
            print('\033[31m\033[1mToken load FAILED.\033[0m')

    ################# Gerar um token #################

    def get_access_token(self):

        url = 'https://api.digikey.com/v1/oauth2/token'

        url_data = {
            'code': self.code,
            'client_id': self.token['client_id'],
            'client_secret': self.token['client_secret'],
            'redirect_uri': 'https://localhost',
            'grant_type': 'authorization_code'
        }

        response = requests.post(url, data=url_data)
        
        
        # se a página entra normalmente
        if response.status_code == 200:
            print('\033[32mAccess Token get SUCCESS\033[0m')

            # alimenta o token
            response_data = response.json()
            self.token['access_token'] = response_data['access_token']
            self.token['refresh_token'] = response_data['refresh_token']
            self.token['expires_in'] = response_data['expires_in']
            self.token['refresh_token_expires_in'] = response_data['refresh_token_expires_in']
            self.token['token_type'] = response_data['token_type']
        # se não
        else: 
            # imprimir o código e qual foi o erro
            print(response)
            print(response.json())

        # guarda o token dentro de uma arquivo .json
        with open(self.filename, "w") as arquivo:
            json.dump(self.token, arquivo)

    ################ Atualizar o token ################

    def get_refresh_token(self):
        url = 'https://api.digikey.com/v1/oauth2/token'

        # Se o token existe
        if self.token == None:
            url_data = {
                'client_id': self.token['client_id'],
                'client_secret': self.token['client_secret'],
                'refresh_token': self.token['refresh_token'],
                'grant_type': 'refresh_token'
            }

            response = requests.post(url, data=url_data)

            # Se entrar no site
            if response.status_code == 200:
                # alimenta o token novemente
                response_data = response.json()
                self.token['access_token'] = response_data['access_token']
                self.token['refresh_token_time'] = time.time()
                self.token['refresh_token'] = response_data['refresh_token']
                self.token['expires_in'] = response_data['expires_in']
                self.token['refresh_token_expires_in'] = response_data['refresh_token_expires_in']
                self.token['token_type'] = response_data['token_type']

                # Verificar se o arquivo existe antes de tentar deletá-lo
                if os.path.exists(self.filename):
                    os.remove(self.filename)

                # cria um novo arquivo .json para armazenar o novo token
                with open(self.filename, "w") as arquivo:
                    json.dump(self.token, arquivo)
            # Se não
            else:
                msg = response.json()

                # Mostra o código e a mensagem de erro
                print('\033[31m\033[1mToken refreshed FAILED\033[0m')
                print(response)
                print(msg['ErrorMessage'])
        # Se não existe
        else:
            print("O token não foi gerado")

    ############### Alimentar o .json ################

    # Pegar as informações da planilha de entrada
    def get_dates_sheet (self):
        # Carregar o arquivo Excel
        workbook = load_workbook(self.sheet)

        # Listar as planilhas disponíveis no arquivo (opcional)
        print(workbook.sheetnames)

        # Escolher uma planilha específica para trabalhar
        sheet = workbook['entrada']

        vetores = []
        self.quants = []

        # Iterar sobre todas as colunas na planilha
        for column in sheet.iter_cols(values_only=True):
            vetores.append(column)

        for a in range(2):
            for i in range(len(vetores[a])):
                if str(vetores[a][i]) != 'None':
                    if a == 0:
                        self.partnumbers.append(vetores[a][i])
                    else:
                        self.quants.append(vetores[a][i])

        for i in range(len(self.partnumbers)):
            self.partnumbers[i] = str(self.partnumbers[i])

        print(self.partnumbers)
        print(self.quants)

    # Pegar informações dos PartNumbers fornecidos
    def get_product_details(self):
        lista = []

        # pesquisa todas as informações dos partnumber fornecidos no arquivo, na DigiKey
        for i in range(len(self.partnumbers)):
            # url do componente
            partnumber_quoted = self.partnumbers[i].replace('/','%2F')
            partnumber_quoted = partnumber_quoted.replace('+','%2B')
            partnumber_quoted = partnumber_quoted.replace('#','%23')
            url = f'https://api.digikey.com/Search/v3/Products/{partnumber_quoted}'

            print(url)
                
            url_header = {
                'x-digikey-locale': 'pt',
                'X-DIGIKEY-Locale-Site': 'BR',
                'X-DIGIKEY-Locale-Currency': 'BRL',
                'Authorization': f"{self.token['token_type']} {self.token['access_token']}",
                'X-DIGIKEY-Client-Id': self.token['client_id']
            }

            response = requests.get(url, headers=url_header)
            
            print(response.url)

            # Se entrou na url
            if response.status_code == 200:
                print(f'\033[32mGot information for {self.partnumbers[i]}\033[0m')
                # alimenta uma lista com um dicionário com o partnumber e a descrição do componente
                if len(response.json()["StandardPricing"]) != 0: 
                    lista.append({'Quantidade': self.quants[i], 'Partnumber': response.json()["ManufacturerPartNumber"], "Description": response.json()["ProductDescription"], "Preco-unitario": response.json()["StandardPricing"][0]["UnitPrice"]})
                else:
                    lista.append({'Quantidade': self.quants[i], 'Partnumber': response.json()["ManufacturerPartNumber"], "Description": response.json()["ProductDescription"], "Preco-unitario": 'null'})
                    print(f'\033[31mTem algo de errado no preço do componente {self.partnumbers[i]}\033[0m')
            # Se não
            else:
                lista.append({'Quantidade': 'null', 'Partnumber': self.partnumbers[i], "Description": 'Esse componente nao foi encontrado', "Preco-unitario": "null"})
                msg = response.json()
                print(f'\033[31mFailed to get information for {self.partnumbers[i]}\033[0m')
                print(response)
                print(msg['ErrorMessage'])

        
        caminho_do_arquivo = 'dados.json'

        # se o arquivo existir
        if os.path.exists(caminho_do_arquivo):
            # deleta o arquivo
            os.remove(caminho_do_arquivo)

        # cria um novo arquivo com as novas informações
        with open("dados.json", "w") as file:
            json.dump(lista, file)    

    ########## Alimentar a planilha de saída ##########

    # Verifica se a tabela está completa
    def check_table (self):
        # Carregar o arquivo Excel
        workbook = load_workbook('planilha.xlsx')

        # Listar as planilhas disponíveis no arquivo (opcional)
        print(workbook.sheetnames)

        # Escolher uma planilha específica para trabalhar
        sheet = workbook['Sheet1']

        for i in range(len(self.data)):
            cell = 'A' + str(i + 2)

            if sheet[cell].value == "null":
                return False

        return True

    #Descobre o preco total da placa
    def financial_table (self):
        # Carregar o arquivo Excel
        workbook = load_workbook('planilha.xlsx')

        # Listar as planilhas disponíveis no arquivo (opcional)
        print(workbook.sheetnames)

        # Escolher uma planilha específica para trabalhar
        sheet = workbook['Sheet1']

        # Calcula o preco total da placa
        preco_placa = 0

        for i in range(len(self.data)):
            quant = 'A' + str(i + 2)
            price = 'D' + str(i + 2)

            if sheet[price].value != 'null':
                preco_placa += int(sheet[quant].value) * float(sheet[price].value)

        return [round(preco_placa, 2), round(5*preco_placa, 2), round(10*preco_placa, 2), round(25*preco_placa, 2), round(50*preco_placa, 2), round(100*preco_placa, 2)]

    # Preenche a planilha de saída
    def filling_out_spreadsheet(self):
        # Criando a planilha
        workbook = xlsxwriter.Workbook('planilha.xlsx')
        worksheet = workbook.add_worksheet()
        
        # Preenchendo os títulos da tabela na planilha
        worksheet.write('A1', 'Quant.')
        worksheet.write('B1', 'PartNumber')
        worksheet.write('C1', 'Description')
        worksheet.write('D1', 'Preço Unitário')

        with open("dados.json", "r") as file:
            self.data = json.load(file)

        # Alimentando a planilha
        for i in range(len(self.data)):
            column1 = 'A' + str(i + 2)
            column2 = 'B' + str(i + 2)
            column3 = 'C' + str(i + 2)
            column4 = 'D' + str(i + 2)
            quant = self.data[i]["Quantidade"]
            partnumber = self.data[i]["Partnumber"]
            description = self.data[i]["Description"]
            preco = self.data[i]["Preco-unitario"]

            worksheet.write(column1, quant)
            worksheet.write(column2, partnumber)
            worksheet.write(column3, description)
            worksheet.write(column4, preco)

        workbook.close()

        # Verificando se a tabela está completa
        ver = self.check_table()
        res = 'y'

        if ver == False:
            print("Sua planilha está incompleta. Deseja continuar? (y/n)")
            res = input()
        
        if res == 'y':
            tabela_financeira = self.financial_table()

        # Fazer a tabela financeira na planilha
        # Carregar o arquivo Excel
        workbook = load_workbook('planilha.xlsx')

        # Listar as planilhas disponíveis no arquivo (opcional)
        print(workbook.sheetnames)

        # Escolher uma planilha específica para trabalhar
        worksheet = workbook['Sheet1']

        worksheet['F1'].value = "Tabela Financeira"
        worksheet['F2'].value = "1"
        worksheet['F3'].value = "5"
        worksheet['F4'].value = "10"
        worksheet['F5'].value = "25"
        worksheet['F6'].value = "50"
        worksheet['F7'].value = "100"

        for i in range(len(tabela_financeira)):
            celula = 'G' + str(i + 2)
            worksheet[celula] = tabela_financeira[i]

        workbook.save('planilha.xlsx')
        
        
    # Estiliza a planilha de saída
    def style_sheet (self):
        # Carregar o arquivo Excel
        workbook = load_workbook('planilha.xlsx')

        # Listar as planilhas disponíveis no arquivo (opcional)
        print(workbook.sheetnames)

        # Escolher uma planilha específica para trabalhar
        sheet = workbook['Sheet1']
        col = 0

        # Muda a largura das colunas
        for column_cells in sheet.columns:
            col+=1

            if col <= 4:
                max_length = 0
                column = column_cells[0].column_letter  # Coluna A, B, C, ...
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        #mescla e centralizar duas celulas
        sheet.merge_cells('F1:G1')

        # Negrito na primeira linha inteira
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Pintar de vermelho os componentes não encontrados
        for i in range(len(self.data)):
            cell1 = 'A' + str(i + 2)
            cell2 = 'B' + str(i + 2)
            cell3 = 'C' + str(i + 2)
            cell4 = 'D' + str(i + 2)

            if sheet[cell1].value == 'null':
                sheet[cell1].font = Font(color="FF0000")
                sheet[cell2].font = Font(color="FF0000")
                sheet[cell3].font = Font(color="FF0000")
                sheet[cell4].font = Font(color="FF0000")
        
        # Defina o estilo da borda
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Defina o alinhamento centralizado
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Aplique a borda e o alinhamento a tabela principal
        for row in sheet.iter_rows(min_row=1, max_col=4, max_row= (len(self.data)+1)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment

        # Aplica a borda e o alinhamento a tabela fianceira
        for row in sheet.iter_rows(min_row=1, min_col=6, max_col=7, max_row= 7):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_alignment
        
        workbook.save('planilha.xlsx')

    ########## Tentando alimentar o programa com as listas da DigiKey no usuário da empresa ############



    def get_list_digi_key (self):
        url = "https://www.digikey.com/MyDigiKey"

        url_header = {
            'x-digikey-locale': 'pt',
            'Authorization': f"{self.token['token_type']} {self.token['access_token']}",
            'X-DIGIKEY-Client-Id': self.token['client_id']
        }
        response = requests.get(url, url_header)

        print(response)
        print(response.json())

# get_list_digi_key()

API()